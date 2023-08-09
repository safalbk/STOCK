import openpyxl


cap=4911
lcap=4*cap



# wb = openpyxl.Workbook()
# sheet = wb.active
# c1 = sheet.cell(row=1, column=2)
# c1.value = 5
# src='/storage/emulated/0/Download/a.xlsx'
# wb.save(src)
# MW-NIFTY-50-15-Nov-2021
import datetime
import os
import glob
import csv
from xlsxwriter.workbook import Workbook
import json
from tabulate import tabulate

x = datetime.datetime.now()
n="MW-NIFTY-50-"
day=x.strftime("%d")
month=x.strftime("%b")
year=x.strftime("%Y")
r=n+day+'-'+month+'-'+year+".csv"
n=n+day+'-'+month+'-'+year
print(r)
# C:\\Users\\HP\\Desktop\\stock\\excel nifty50
path=""
# /storage/emulated/0/stock/File
# /storage/emulated/0/Download/bk_report-converted (1).pdf
src='C:\\Users\\stark\\Downloads'
tar="C:\\Users\\stark\\Desktop\\stock\\excel nifty50"
# src='C:\Users\stark\Downloads'
# tar="/storage/emulated/0/stock"
csvfile2 = glob.glob(os.path.join(src, r))
for csvfile in glob.glob(os.path.join(src, r)):
    # print(csvfile[22:-4])
    workbook = Workbook(os.path.join(tar, n)
 + '.xlsx')
    path =(os.path.join(tar, n)
 + '.xlsx')
    worksheet = workbook.add_worksheet()
    with open(csvfile, 'rt', encoding='utf8') as f:
        reader = csv.reader(f)
        worksheet.write(0, 0, "SYMBOL")
        worksheet.write(0, 1, "OPEN")
        worksheet.write(0, 2, "HIGH")
        worksheet.write(0, 3, "LOW")
        worksheet.write(0, 4, "PREV. CLOSE")
        worksheet.write(0, 5, "LTP")
        worksheet.write(0, 6, "CHNG")
        worksheet.write(0, 7, "%CHNG")
        worksheet.write(0, 8, "VOLUME")
        worksheet.write(0, 9, "VALUE")
        worksheet.write(0, 10, "52W H ")
        worksheet.write(0, 11, "52W L")
        worksheet.write(0, 12, "365 D % CHNG")
        worksheet.write(0, 13, "30 D % CHNG")

        for r, row in enumerate(reader):
            # print(r)
            for c, col in enumerate(row):
                # print(r," ",col)
                if(r>9):
                    worksheet.write(r-9, c, col)
    workbook.close()
# os.replace(csvfile2[0],os.path.join('C:\\Users\\HP\\Desktop\\stock\\csv nifty50', csvfile2[0][22:-4]+'.csv'))
def find_sector(symbol):
    sector = ""
    with open("nifty50.json","r") as f:
        n50_data = json.loads(f.read())
        n50_data["Nifty50"][symbol]["sector"]
        sector = n50_data["Nifty50"][symbol]["sector"]
    return str(sector)

def find_points():
    pass
# print(find_sector("INFY"))
# input()
# exit()
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
print("   open = high (sell side)")
data =[]
for i in range(2,sheet_obj.max_row):

    open_value = sheet_obj.cell(row=i, column=2)
    open_value=open_value.value
    high = sheet_obj.cell(row=i, column=3)
    high=high.value
    name = sheet_obj.cell(row=i, column=1)
    name=name.value
    ltp = sheet_obj.cell(row=i, column=6)
    ltp=ltp.value

    if(open_value==high and open_value!=""):
        sector = find_sector(name)
        row_data = [name, ltp, int(lcap / float(ltp.replace(',', ''))), sector]
        data.append(row_data)
        # print("   ",name,"  ", ltp,"      stocks :",int(lcap/float(ltp.replace(',','')))," Sector :",find_sector(name))
headers = ["Name", "LTP", "Stocks", "Sector"]
print(tabulate(data, headers=headers, tablefmt="grid"))

print("")
data =[]
print("   open = low (buy side)")
for i in range(2,sheet_obj.max_row):
    open_value = sheet_obj.cell(row=i, column=2)
    open_value=open_value.value
    low = sheet_obj.cell(row=i, column=4)
    low=low.value
    name = sheet_obj.cell(row=i, column=1)
    name=name.value
    ltp = sheet_obj.cell(row=i, column=6)
    ltp=ltp.value
    if(open_value==low  and open_value!=""):
        sector = find_sector(name)
        row_data = [name, ltp, int(lcap / float(ltp.replace(',', ''))), sector]
        data.append(row_data)
        # print("   ",name,"  ", ltp,"      stocks :",int(lcap/float(ltp.replace(',','')))," Sector :",find_sector(name))
headers = ["Name", "LTP", "Stocks", "Sector"]
print(tabulate(data, headers=headers, tablefmt="grid"))

a=input("enter to exits")
